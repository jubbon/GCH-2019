#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
import json
import re
import time

from pathlib import Path

from openpyxl import load_workbook

import geocoder


def get_coordinates(address, prefix):
    '''
    Запрашивает координаты дома по его адресу
    '''
    try:
        if prefix not in address:
            address = prefix + ", " + address
        print(f"[get_coordinates] Getting coordinates for '{address}'",flush=True)
        data = geocoder.yandex(location=address, lang="ru-RU", kind="house")
        geo = data.json
        assert isinstance(geo, dict)
        return dict(
            longitude=float(geo["lng"]),
            latitude=float(geo["lat"])
        )
    except Exception as err:
        print(f"[get_coordinates] Error: {err}", flush=True)


def normalize_address(address):
    ''' Выполняет нормализацию адреса
    '''
    parts = {}
    for address_part in address.split(","):
        address_part = address_part.strip()
        if re.fullmatch("\\d{6}", address_part):
            parts["postal_index"] = address_part
        elif address_part.startswith("обл "):
            parts["region"] = {
                "type": "область",
                "value": address_part.replace("обл ", "")
            }
        elif address_part.startswith("край "):
            parts["region"] = {
                "type": "край",
                "value": address_part.replace("край ", "")
            }
        elif address_part.startswith("г "):
            parts["city"] = {
                "type": "город",
                "value": address_part.replace("г ", "")
            }
        elif address_part.startswith("п ЗАТО п. "):
            parts["city"] = {
                "type": "поселок",
                "value": address_part.replace("п ЗАТО п. ", "")
            }
        elif address_part.startswith("ул "):
            parts["street"] = {
                "type": "улица",
                "value": address_part.replace("ул ", "")
            }
        elif address_part.startswith("пр-кт "):
            parts["street"] = {
                "type": "проспект",
                "value": address_part.replace("пр-кт ", "")
            }
        elif address_part.startswith("проезд "):
            parts["street"] = {
                "type": "проезд",
                "value": address_part.replace("проезд ", "")
            }
        elif address_part.startswith("пл "):
            parts["street"] = {
                "type": "площадь",
                "value": address_part.replace("пл ", "")
            }
        elif address_part.startswith("пер "):
            parts["street"] = {
                "type": "переулок",
                "value": address_part.replace("пер ", "")
            }
        elif address_part.startswith("д. "):
            parts["house"] = {
                "type": "дом",
                "value": address_part.replace("д. ", "")
            }
        elif address_part.startswith("к. "):
            parts["building"] = {
                "type": "корпус",
                "value": address_part.replace("к. ", "")
            }
        elif address_part.startswith("стр. "):
            parts["building"] = {
                "type": "строение",
                "value": address_part.replace("стр. ", "")
            }
        else:
            raise RuntimeError(
                f"Unknown address part '{address_part}' for '{address}'")
    return parts


def geocode_address(address):
    ''' Выполняет геокодирование адреса
    '''
    geo = geocoder.yandex(location=address, lang="ru-RU", kind="house")
    data = geo.json
    assert isinstance(data, dict)
    data.update(provider="yandex")
    return data


def handle_dom(meta, data):
    '''
    '''
    assert isinstance(data, dict)
    region_uid = meta["region"]

    print(f"Received {len(data)} addresses", flush=True)
    for _, house in data.items():
        address = house["address"]
        address_fullname = address["fullname"]

        # Нормализация адреса
        if "parts" not in address:
            try:
                print(f"Normalizing address {address_fullname}...", flush=True)
                address_parts = normalize_address(address_fullname)
                data = {
                    "fullname": address_fullname,
                    "parts": address_parts
                }
            except Exception as err:
                print(f"[ERROR] {err}", flush=True)

        # Геокодирование адреса
        if "geo" not in house:
            try:
                print(f"Geocoding address {address_fullname}...", flush=True)
                address_geo = geocode_address(address_fullname)
                data = {
                    "fullname": address_fullname,
                    "geo": address_geo
                }
            except Exception as err:
                print(f"[ERROR] {err}", flush=True)


def geocode_dumpsters(filename, force=False):
    '''
    '''
    wb = load_workbook(filename=filename)
    for sheet_name in ("4.1.1.1 Автозаводский", "4.1.1.2 Канавинский",
                       "4.1.1.3 Ленинский", "4.1.1.4 Московский",
                       "4.1.1.5 Нижегородский", "4.1.1.6 Приокский",
                       "4.1.1.7 Советский", "4.1.1.8 Сормовский"):
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_col=4, max_col=4, min_row=8):
            cell = row[0]
            address = cell.value
            row_number = cell.row
            coords = ws[f'N{row_number}'].value
            if address and (force or not coords):
                coords = get_coordinates(
                    address.strip(), prefix="Нижний Новгород")
                if coords:
                    latitude = round(coords["latitude"], 5)
                    longitude = round(coords["longitude"], 5)
                    ws[f'N{row_number}'] = f"{latitude}, {longitude}"
                else:
                    ws[f'N{row_number}'] = f"Координаты не заданы"
            else:
                print(f"Skipping for {address}", flush=True)
    wb.save(filename)


def main():
    '''
    '''
    force = False
    # geocode_dumpsters(
    #     filename="/data/Реестр контейнерных площадок.xlsx",
    #     force=force
    # )
    geocode_houses(
        input_filename="/data/Реестр жилых домов.json",
        output_filename="/data/Координаты жилых домов.json",
        force=force
    )
