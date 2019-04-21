#!/usr/bin/env python
# -*- coding: utf-8 -*-

import json
from pathlib import Path

import requests
from openpyxl import load_workbook


def load_dumpsters(filename):
    '''
    '''
    dumpsters = dict()
    wb = load_workbook(filename=filename)
    for sheet_name in ("4.1.1.1 Автозаводский", "4.1.1.2 Канавинский",
                       "4.1.1.3 Ленинский", "4.1.1.4 Московский",
                       "4.1.1.5 Нижегородский", "4.1.1.6 Приокский",
                       "4.1.1.7 Советский", "4.1.1.8 Сормовский"):
        ws = wb[sheet_name]
        area_name = sheet_name.split()[-1]
        for row in ws.iter_rows(min_col=1, max_col=14, min_row=8):
            dumpster_number = row[0].value
            dumpster = {
                "Округ": area_name,
                "Номер контейнерной площадки": row[1].value,
                "Адрес": row[3].value,
                "Наименование": row[4].value,
                "Владелец": row[5].value,
                "Транспортировщик": row[6].value,
                "Материал": row[7].value,
                "Количество": row[8].value,
                "Вместимомть": row[9].value,
                "Покрытие": row[10].value,
                "Навес": row[11].value,
                "Координаты дома": row[13].value
            }
            dumpsters[f"{area_name}_{dumpster_number}"] = dumpster
    return dumpsters


def link_to_road(dumpsters):
    '''
    '''
    for dumpster_uid, dumpster in dumpsters.items():
        coordinates = dumpster.get("Координаты дома", "")
        if not coordinates:
            print(f"Нет координат для: {dumpster_uid}", flush=True)
            continue
        coordinates = ",".join(coordinates.split(", ")[::-1])
        url = f"http://osrm.vehicle:5000/nearest/v1/car/{coordinates}"

        response = requests.get(url)
        nearest_point = response.json()
        status = nearest_point["code"]
        if status == "Ok":
            points = nearest_point.get("waypoints")
            dumpster.update({"Расположение": points})
        else:
            print(nearest_point, flush=True)
    return dumpsters


def save_to_json(output_filename, dumpsters):
    '''
    '''
    assert isinstance(dumpsters, dict)
    data = json.dumps(dumpsters, ensure_ascii=False, indent=4)
    Path(output_filename).write_text(data, encoding="utf-8")


def load_from_json(filename):
    ''' Загружает данные из JSON-файла
    '''
    data = json.loads(Path(filename).read_text())
    assert isinstance(data, dict)
    return data


def link_to_nearest_dumpster(coords, dumpsters):
    '''
    '''
    nearest_dumpsters = dict()
    total = 0
    for address, coord in coords.items():
        # Контейнерные площадки, доступные для дома
        available_dumpsters = list()
        for coord_threshold in (0.005, 0.01, 0.015, 0.02, 0.03, 0.04, 0.05, 0.1):
            latitude = float(coord["lat"])
            longitude = float(coord["lng"])

            for dumpster_uid, dumpster in dumpsters.items():
                try:
                    dumster_location = dumpster["Расположение"][0]["location"]
                    dumpster_longitude, dumpster_latitude = dumster_location
                    if dumpster_latitude > latitude + coord_threshold:
                        continue
                    if dumpster_latitude < latitude - coord_threshold:
                        continue
                    if dumpster_longitude > longitude + coord_threshold:
                        continue
                    if dumpster_longitude < longitude - coord_threshold:
                        continue
                except KeyError:
                    # print(f"Нет координат для площадки {dumpster_uid}", flush=True)
                    continue
                dumpster.update(uid=dumpster_uid)
                available_dumpsters.append(dumpster)
            if len(available_dumpsters) >= 3:
                break

        dumpster_coords = list()
        for dumpster in available_dumpsters:
            dumster_location = dumpster["Расположение"][0]["location"]
            dumpster_longitude, dumpster_latitude = dumster_location
            dumpster_coords.append(f"{dumpster_longitude},{dumpster_latitude}")
        coordinates = ";".join([f"{longitude},{latitude}"] + dumpster_coords)
        
        url = f"http://osrm.human:5000/table/v1/foot/{coordinates}"
        params = dict(
            sources="0",
            destinations=";".join([str(n) for n, d in enumerate(available_dumpsters, 1)])
        )
        min_duration = None

        response = requests.get(url, params=params)
        data = response.json()
        status = data["code"]
        if status == "Ok":
            durations = data.get("durations")[0]
            for n, duration in enumerate(durations):
                if min_duration is None or duration < min_duration[1]:
                    min_duration = (n, duration)
        else:
            print(data, flush=True)

        if min_duration:
            dumpster_index, duration_value = min_duration
            dumpster = available_dumpsters[dumpster_index]
            dumpster.update(duration=duration_value)
            nearest_dumpsters[address] = dumpster

        total += 1
        if total % 100 == 0:
            print(f"Обработано домов: {total}", flush=True)
    return nearest_dumpsters


def main():
    '''
    '''
    dumpsters = load_dumpsters(filename="/data/Реестр контейнерных площадок.xlsx")
    # Привязка контейнеров к дорогам
    dumpsters = link_to_road(dumpsters)
    save_to_json("/data/Реестр контейнерных площадок.json", dumpsters)
    print(len(dumpsters), flush=True)

    # Првязка домов к контейнерным площадкам
    coords = load_from_json("/data/Координаты жилых домов.json")
    nearest_dumpsters = link_to_nearest_dumpster(coords, dumpsters)
    save_to_json("/data/Ближайшие контейнерные площадки.json", nearest_dumpsters)
